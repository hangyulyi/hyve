import openpyxl
import json

with open('data.json', 'r', encoding='utf-8') as file:
    data_list = json.load(file)

directory = 'data'

wb = openpyxl.load_workbook('enter excel file here.xlsx')

# Will work with multiple sheets
for sheet in wb.worksheets:
    # adjust row/col accordingly, will grab from two different columns and all the rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=8, values_only=True):
        # base row[n] off of max col wanted ie. row[0] with min_col = 6 will make kr = column 6
        kr = row[0]
        eng = row[2]
        if kr is not None and eng is not None:
            # check that all cells aren't empty
            data_list.append({
                'kr': kr,
                'eng': eng
            })
            


with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data_list, f, ensure_ascii=False, indent=4)


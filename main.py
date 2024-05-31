import openpyxl

dataframe = openpyxl.load_workbook("leniverse\[LENIVERSE] EP_[LENIVERSE] EP.19 나랑 하이브 같이 가자! 1편ENG_평가완료.xlsx")

dataframe1 = dataframe.active

for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
import openpyxl
import json
import os


data_list = []

# insert name of folder you wish to iterate through with excel files
directory = ''

for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)

        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # adjust row/col accordingly, will grab from two different columns and all the rows
        # Can refer to xlsx-to-json-copy.py file for more comments if needed
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=4, values_only=True):
            data_list.append({
                'kr': row[0],
                'eng': row[1]
            })

json_data = json.dumps(data_list, ensure_ascii=False, indent=4)

with open('data.json', 'w', encoding='utf-8') as f:
    f.write(json_data)

